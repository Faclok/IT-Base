import base64
import csv
import datetime as dt
import hashlib
import hmac
import io
import json
import math
import os
import re
import time
from collections import Counter
from pathlib import Path
from typing import Literal

import aiosqlite
import httpx
from fastapi import FastAPI, File, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
WEB_DIR = BASE_DIR / "web"
DB_PATH = os.getenv("IT_BASE_DB_PATH", str(BASE_DIR / "backend" / "itbase.sqlite3"))
ADMIN_PASSWORD = os.getenv("IT_BASE_ADMIN_PASSWORD", "change_me")
COOKIE_SECRET = os.getenv("IT_BASE_COOKIE_SECRET", ADMIN_PASSWORD).encode()
COOKIE_NAME = "itbase_admin"
COOKIE_TTL = 60 * 60 * 12
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
GRADE_ORDER = {"Junior": 0, "Middle": 1, "Senior": 2, "Lead": 3}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TELEGRAM_RE = re.compile(r"^@[\w\d_]{3,}$")


class AdminLoginIn(BaseModel):
    password: str


class DeveloperIn(BaseModel):
    name: str
    title: str
    stack: str = ""
    skills: list[str] = Field(default_factory=list)
    experience: str = ""
    grade: Literal["Junior", "Middle", "Senior", "Lead"]
    contact_email: str = ""
    contact_telegram: str = ""


class ContactRequestIn(BaseModel):
    developer_id: int
    customer_telegram: str = Field(pattern=r"^@[\w\d_]{3,}$")
    message: str = ""


class ContactRequestStatusIn(BaseModel):
    status: Literal["new", "processed"]


def n(v: str) -> str:
    return re.sub(r"\s+", " ", (v or "")).strip()


def validate_developer_payload(payload: DeveloperIn) -> DeveloperIn:
    payload.name = n(payload.name)
    payload.title = n(payload.title)
    payload.stack = n(payload.stack)
    payload.experience = n(payload.experience)
    payload.contact_email = n(payload.contact_email)
    payload.contact_telegram = n(payload.contact_telegram)
    payload.skills = [n(x) for x in payload.skills if n(x)]

    if len(payload.name) < 2:
        raise HTTPException(status_code=400, detail="name_too_short")
    if len(payload.title) < 2:
        raise HTTPException(status_code=400, detail="title_too_short")
    if payload.contact_email and not EMAIL_RE.match(payload.contact_email):
        raise HTTPException(status_code=400, detail="bad_contact_email")
    if payload.contact_telegram and not TELEGRAM_RE.match(payload.contact_telegram):
        raise HTTPException(status_code=400, detail="bad_contact_telegram")
    if len(payload.skills) > 50:
        raise HTTPException(status_code=400, detail="too_many_skills")
    return payload


HEADER_ALIASES = {
    "id": "source_id",
    "name": "name",
    "fio": "name",
    "nickname": "name",
    "ник": "name",
    "фио": "name",
    "title": "title",
    "role": "title",
    "position": "title",
    "позиция": "title",
    "stack": "stack",
    "стек": "stack",
    "skills": "skills",
    "навыки": "skills",
    "experience": "experience",
    "опыт": "experience",
    "grade": "grade",
    "грейд": "grade",
    "contact_email": "contact_email",
    "email": "contact_email",
    "почта": "contact_email",
    "contact_telegram": "contact_telegram",
    "telegram": "contact_telegram",
    "телеграм": "contact_telegram",
}


def parse_skills(raw: str) -> list[str]:
    return [n(x) for x in re.split(r"[,\n;/|]+", raw or "") if n(x)]


def map_row(raw: dict) -> DeveloperIn:
    normalized = {}
    for key, value in raw.items():
        out_key = HEADER_ALIASES.get(n(str(key)).lower())
        if out_key:
            normalized[out_key] = n(str(value or ""))
    source_id = normalized.get("source_id", "")
    auto_name = f"Candidate {source_id}" if source_id else "Candidate Imported"
    name_value = normalized.get("name", "") or auto_name
    title_value = normalized.get("title", "") or "Developer"
    grade_value = normalized.get("grade", "") or "Junior"
    payload = DeveloperIn(
        name=name_value,
        title=title_value,
        stack=normalized.get("stack", ""),
        skills=parse_skills(normalized.get("skills", "")),
        experience=normalized.get("experience", ""),
        grade=grade_value,
        contact_email=normalized.get("contact_email", ""),
        contact_telegram=normalized.get("contact_telegram", ""),
    )
    return validate_developer_payload(payload)


def sign(ts: int) -> str:
    raw = str(ts).encode()
    sig = hmac.new(COOKIE_SECRET, raw, hashlib.sha256).digest()
    return f"{ts}.{base64.urlsafe_b64encode(sig).decode().rstrip('=')}"


def verify(v: str) -> bool:
    try:
        ts_s, s = v.split(".", 1)
        if time.time() - int(ts_s) > COOKIE_TTL:
            return False
        pad = "=" * ((4 - len(s) % 4) % 4)
        got = base64.urlsafe_b64decode((s + pad).encode())
        exp = hmac.new(COOKIE_SECRET, ts_s.encode(), hashlib.sha256).digest()
        return hmac.compare_digest(got, exp)
    except Exception:
        return False


async def req_admin(request: Request):
    if not verify(request.cookies.get(COOKIE_NAME, "")):
        raise HTTPException(status_code=401, detail="admin_required")


def tok(t: str) -> list[str]:
    return re.findall(r"[a-zA-Zа-яА-Я0-9\+\#]{2,}", (t or "").lower())


def rank_tfidf(query: str, rows: list[aiosqlite.Row]) -> list[dict]:
    docs = []
    for r in rows:
        docs.append(tok(" ".join([r["name"], r["title"], r["stack"], r["experience"], " ".join(json.loads(r["skills_json"]))])))
    q = tok(query)
    if not q:
        return []
    df = Counter()
    for d in docs:
        for t in set(d):
            df[t] += 1
    n_docs = max(1, len(docs))
    qtf = Counter(q)
    qv = {t: c * (math.log((n_docs + 1) / (df.get(t, 0) + 1)) + 1.0) for t, c in qtf.items()}
    qn = math.sqrt(sum(v * v for v in qv.values())) or 1.0
    scored = []
    for i, d in enumerate(docs):
        dtf = Counter(d)
        dv = {t: c * (math.log((n_docs + 1) / (df.get(t, 0) + 1)) + 1.0) for t, c in dtf.items()}
        dot = sum(qv.get(t, 0) * dv.get(t, 0) for t in qv)
        dn = math.sqrt(sum(v * v for v in dv.values())) or 1.0
        scored.append((dot / (qn * dn), rows[i]))
    scored.sort(key=lambda x: x[0], reverse=True)
    out = []
    for s, r in scored[:5]:
        if s <= 0:
            continue
        out.append(to_public(r) | {"score": round(float(s), 4)})
    return out


def to_public(r: aiosqlite.Row) -> dict:
    return {
        "id": r["id"],
        "created_at": r["created_at"],
        "name": r["name"],
        "title": r["title"],
        "stack": r["stack"] or "",
        "skills": json.loads(r["skills_json"] or "[]"),
        "experience": r["experience"] or "",
        "grade": r["grade"],
    }


app = FastAPI(title="IT Base")
app.mount("/static", StaticFiles(directory=str(WEB_DIR)), name="web")


@app.on_event("startup")
async def startup():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("pragma foreign_keys = on")
        await db.executescript(
            """
            create table if not exists developers (
              id integer primary key autoincrement,
              created_at text not null default (datetime('now')),
              name text not null, title text not null, stack text not null default '',
              skills_json text not null default '[]', experience text not null default '',
              grade text not null, contact_email text not null default '', contact_telegram text not null default ''
            );
            create table if not exists contact_requests (
              id integer primary key autoincrement,
              created_at text not null default (datetime('now')),
              developer_id integer not null references developers(id) on delete cascade,
              customer_telegram text not null, message text not null default '', status text not null default 'new'
            );
            """
        )
        await db.commit()


@app.get("/")
async def index():
    return FileResponse(WEB_DIR / "index.html")


@app.get("/admin")
async def admin_page():
    return FileResponse(WEB_DIR / "admin.html")


@app.get("/api/admin/me")
async def me(request: Request):
    return {"isAdmin": verify(request.cookies.get(COOKIE_NAME, ""))}


@app.post("/api/admin/login")
async def login(payload: AdminLoginIn, response: Response):
    if payload.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="bad_password")
    response.set_cookie(COOKIE_NAME, sign(int(time.time())), httponly=True, samesite="lax", max_age=COOKIE_TTL, path="/")
    return {"ok": True}


@app.post("/api/admin/logout")
async def logout(response: Response):
    response.delete_cookie(COOKIE_NAME, path="/")
    return {"ok": True}


@app.get("/api/admin/backup")
async def backup_db(request: Request):
    await req_admin(request)
    source = Path(DB_PATH)
    if not source.exists():
        raise HTTPException(status_code=404, detail="db_not_found")
    backup_dir = source.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    target = backup_dir / f"itbase-backup-{stamp}.sqlite3"
    target.write_bytes(source.read_bytes())
    return FileResponse(path=target, filename=target.name, media_type="application/octet-stream")


@app.get("/api/public/developers")
async def list_public(query: str = "", grade: str = "", sort: str = "date_desc", page: int = 1, page_size: int = Query(24, ge=1, le=100)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute("select id, created_at, name, title, stack, skills_json, experience, grade from developers")).fetchall()
    q = n(query).lower()
    out = []
    for r in rows:
        if grade and r["grade"] != grade:
            continue
        if q:
            blob = " ".join([r["name"], r["title"], r["stack"], r["experience"], " ".join(json.loads(r["skills_json"] or "[]"))]).lower()
            if q not in blob:
                continue
        out.append(r)
    if sort == "grade_desc":
        out.sort(key=lambda r: (GRADE_ORDER[r["grade"]], r["id"]), reverse=True)
    elif sort == "grade_asc":
        out.sort(key=lambda r: (GRADE_ORDER[r["grade"]], -r["id"]))
    else:
        out.sort(key=lambda r: r["id"], reverse=True)
    total = len(out)
    start = (page - 1) * page_size
    return {"items": [to_public(x) for x in out[start : start + page_size]], "total": total, "page": page, "page_size": page_size}


@app.post("/api/public/recommendations")
async def rec(payload: dict):
    query = n(str(payload.get("query", "")))
    if not query:
        raise HTTPException(status_code=400, detail="query_required")
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute("select id, created_at, name, title, stack, skills_json, experience, grade from developers order by id desc limit 1000")).fetchall()
    if OPENAI_API_KEY:
        try:
            prompt = (
                "Return only JSON array with up to 5 objects "
                '[{"id": number, "score": number, "reason": string}]. '
                f"Task: {query}\nDevelopers:{json.dumps([to_public(r) for r in rows], ensure_ascii=False)}"
            )
            async with httpx.AsyncClient(timeout=20) as c:
                d = await c.post(
                    "https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
                    json={"model": OPENAI_MODEL, "messages": [{"role": "user", "content": prompt}], "temperature": 0.1},
                )
            content = d.json()["choices"][0]["message"]["content"]
            arr = json.loads(content)
            by_id = {r["id"]: to_public(r) for r in rows}
            items = []
            for x in arr[:5]:
                if x.get("id") in by_id:
                    items.append(by_id[x["id"]] | {"score": float(x.get("score", 0)), "reason": str(x.get("reason", ""))})
            if items:
                return {"provider": "openai", "items": items}
        except Exception:
            pass
    return {"provider": "tfidf", "items": rank_tfidf(query, rows)}


@app.get("/api/admin/developers")
async def list_admin(request: Request):
    await req_admin(request)
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (
            await db.execute(
                "select id, created_at, name, title, stack, skills_json, experience, grade, contact_email, contact_telegram from developers order by id desc"
            )
        ).fetchall()
    return [to_public(r) | {"contact_email": r["contact_email"], "contact_telegram": r["contact_telegram"]} for r in rows]


@app.post("/api/admin/developers")
async def create_admin(request: Request, payload: DeveloperIn):
    await req_admin(request)
    payload = validate_developer_payload(payload)
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "insert into developers(name,title,stack,skills_json,experience,grade,contact_email,contact_telegram) values (?,?,?,?,?,?,?,?)",
            (
                n(payload.name),
                n(payload.title),
                n(payload.stack),
                json.dumps([n(x) for x in payload.skills if n(x)], ensure_ascii=False),
                n(payload.experience),
                payload.grade,
                n(payload.contact_email),
                n(payload.contact_telegram),
            ),
        )
        await db.commit()
    return {"ok": True, "id": cur.lastrowid}


@app.post("/api/admin/developers/import")
async def import_admin_developers(request: Request, file: UploadFile = File(...)):
    await req_admin(request)
    filename = (file.filename or "").lower()
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="empty_file")

    rows: list[dict] = []
    if filename.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = [dict(r) for r in reader]
    elif filename.endswith(".xlsx"):
        wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        header = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            values = [n(str(x or "")) for x in row]
            if idx == 0:
                header = values
                continue
            if not any(values):
                continue
            rows.append({header[i]: values[i] if i < len(values) else "" for i in range(len(header))})
    else:
        raise HTTPException(status_code=400, detail="unsupported_file_type")

    if not rows:
        raise HTTPException(status_code=400, detail="no_rows_found")

    ok_payloads: list[DeveloperIn] = []
    errors: list[dict] = []
    for i, row in enumerate(rows, start=2):
        try:
            ok_payloads.append(map_row(row))
        except Exception as e:
            detail = getattr(e, "detail", str(e))
            errors.append({"row": i, "error": str(detail)})

    inserted = 0
    if ok_payloads:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.executemany(
                "insert into developers(name,title,stack,skills_json,experience,grade,contact_email,contact_telegram) values (?,?,?,?,?,?,?,?)",
                [
                    (
                        p.name,
                        p.title,
                        p.stack,
                        json.dumps(p.skills, ensure_ascii=False),
                        p.experience,
                        p.grade,
                        p.contact_email,
                        p.contact_telegram,
                    )
                    for p in ok_payloads
                ],
            )
            await db.commit()
            inserted = len(ok_payloads)

    return {"ok": True, "inserted": inserted, "errors": errors[:30], "total_rows": len(rows)}


@app.put("/api/admin/developers/{developer_id}")
async def update_admin(request: Request, developer_id: int, payload: DeveloperIn):
    await req_admin(request)
    payload = validate_developer_payload(payload)
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "update developers set name=?,title=?,stack=?,skills_json=?,experience=?,grade=?,contact_email=?,contact_telegram=? where id=?",
            (
                n(payload.name),
                n(payload.title),
                n(payload.stack),
                json.dumps([n(x) for x in payload.skills if n(x)], ensure_ascii=False),
                n(payload.experience),
                payload.grade,
                n(payload.contact_email),
                n(payload.contact_telegram),
                developer_id,
            ),
        )
        await db.commit()
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="developer_not_found")
    return {"ok": True}


@app.delete("/api/admin/developers/{developer_id}")
async def delete_admin(request: Request, developer_id: int):
    await req_admin(request)
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("delete from developers where id = ?", (developer_id,))
        await db.commit()
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="developer_not_found")
    return {"ok": True}


@app.post("/api/public/contact-requests")
async def create_req(payload: ContactRequestIn):
    async with aiosqlite.connect(DB_PATH) as db:
        dev = await (await db.execute("select id from developers where id = ?", (payload.developer_id,))).fetchone()
        if not dev:
            raise HTTPException(status_code=404, detail="developer_not_found")
        cur = await db.execute(
            "insert into contact_requests(developer_id, customer_telegram, message, status) values (?, ?, ?, 'new')",
            (payload.developer_id, n(payload.customer_telegram), n(payload.message)),
        )
        await db.commit()
    return {"ok": True, "id": cur.lastrowid}


@app.get("/api/admin/contact-requests")
async def list_reqs(
    request: Request,
    status: Literal["all", "new", "processed"] = "all",
    customer_telegram: str = "",
):
    await req_admin(request)
    filters = []
    params: list[str] = []
    if status != "all":
        filters.append("cr.status = ?")
        params.append(status)
    tg_query = n(customer_telegram).lower()
    if tg_query:
        filters.append("lower(cr.customer_telegram) like ?")
        params.append(f"%{tg_query}%")
    where_sql = f"where {' and '.join(filters)}" if filters else ""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (
            await db.execute(
                """
                select cr.id, cr.created_at, cr.developer_id, cr.customer_telegram, cr.message, cr.status,
                       d.name as developer_name, d.grade as developer_grade
                from contact_requests cr join developers d on d.id = cr.developer_id
                """
                + where_sql
                + """
                order by cr.id desc
                """,
                params,
            )
        ).fetchall()
    return [dict(r) for r in rows]


@app.put("/api/admin/contact-requests/{request_id}/status")
async def req_status(request: Request, request_id: int, payload: ContactRequestStatusIn):
    await req_admin(request)
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("update contact_requests set status = ? where id = ?", (payload.status, request_id))
        await db.commit()
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="request_not_found")
    return {"ok": True}
